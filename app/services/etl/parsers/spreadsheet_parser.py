"""Spreadsheet parser for `.xlsx` workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Mapping
import logging
import posixpath
import uuid
import xml.etree.ElementTree as ET
import zipfile

from app.services.etl.models import ExtractedSegment, ParsedDocument
from app.services.etl.utils import normalize_text_block

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - fallback path only used when openpyxl is unavailable.
    load_workbook = None


_SHEET_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_REL_NS = {"r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
_PKG_REL_NS = {"pr": "http://schemas.openxmlformats.org/package/2006/relationships"}
logger = logging.getLogger(__name__)


class SpreadsheetDocumentParser:
    """Parse `.xlsx` workbooks into sheet-level ETL segments."""

    parser_name = "spreadsheet_document_parser"
    supported_extensions = {".xlsx"}

    def parse(
        self,
        *,
        file_path: Path,
        relative_path: str,
        extra_metadata: Mapping[str, object] | None = None,
    ) -> ParsedDocument:
        """Parse one `.xlsx` workbook."""
        sheets = self._extract_sheets(file_path)
        normalized_sheets = [(name, rows) for name, rows in sheets if rows]
        if not normalized_sheets:
            raise ValueError(f"Document is empty after parsing: {file_path.name}")

        source_id = uuid.uuid4().hex
        segments: list[ExtractedSegment] = []
        markdown_sections = [f"# {file_path.name}"]

        for sequence, (sheet_name, rows) in enumerate(normalized_sheets, start=1):
            plain_text = "\n".join(" | ".join(row) for row in rows if any(cell for cell in row)).strip()
            markdown_body = _rows_to_markdown(rows)
            markdown_text = f"## Sheet: {sheet_name}\n\n{markdown_body}".strip()
            markdown_sections.append(markdown_text)
            segments.append(
                ExtractedSegment(
                    segment_id=uuid.uuid4().hex,
                    sequence=sequence,
                    content_text=plain_text,
                    content_markdown=markdown_text,
                    metadata={
                        "segment_kind": "worksheet",
                        "source_type": "xlsx",
                        "sheet_name": sheet_name,
                    },
                )
            )

        content_json = {
            "source_id": source_id,
            "file_name": file_path.name,
            "file_type": "xlsx",
            "parser_name": self.parser_name,
            "sheet_count": len(segments),
            "segments": [segment.to_dict() for segment in segments],
        }
        return ParsedDocument(
            source_id=source_id,
            source_path=str(file_path),
            relative_path=relative_path,
            file_name=file_path.name,
            file_type="xlsx",
            parser_name=self.parser_name,
            content_markdown="\n\n".join(markdown_sections).strip(),
            content_json=content_json,
            segments=segments,
            metadata=dict(extra_metadata or {}),
        )

    def _extract_sheets(self, file_path: Path) -> list[tuple[str, list[list[str]]]]:
        if load_workbook is not None:
            try:
                return self._extract_with_openpyxl(file_path)
            except Exception:
                logger.exception("openpyxl parsing failed for %s, falling back to zip/xml.", file_path)
        return self._extract_with_zip(file_path)

    def _extract_with_openpyxl(self, file_path: Path) -> list[tuple[str, list[list[str]]]]:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheets: list[tuple[str, list[list[str]]]] = []
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                values = [normalize_text_block("" if value is None else str(value)) for value in row]
                if any(values):
                    rows.append(values)
            if rows:
                sheets.append((worksheet.title, rows))
        workbook.close()
        return sheets

    def _extract_with_zip(self, file_path: Path) -> list[tuple[str, list[list[str]]]]:
        with zipfile.ZipFile(file_path) as archive:
            shared_strings = _load_shared_strings(archive)
            sheet_targets = _load_sheet_targets(archive)
            sheets: list[tuple[str, list[list[str]]]] = []
            for sheet_name, target in sheet_targets:
                with archive.open(target) as handle:
                    sheet_root = ET.fromstring(handle.read())
                rows = _load_sheet_rows(sheet_root, shared_strings)
                if rows:
                    sheets.append((sheet_name, rows))
            return sheets


def _load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    shared_strings: list[str] = []
    for item in root.findall("a:si", _SHEET_NS):
        text = normalize_text_block("".join(node.text or "" for node in item.findall(".//a:t", _SHEET_NS)))
        shared_strings.append(text)
    return shared_strings


def _load_sheet_targets(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rel_root.findall("pr:Relationship", _PKG_REL_NS)
        if rel.attrib.get("Id") and rel.attrib.get("Target")
    }

    sheets: list[tuple[str, str]] = []
    for sheet in workbook_root.findall("a:sheets/a:sheet", _SHEET_NS):
        rel_id = sheet.attrib.get(f"{{{_REL_NS['r']}}}id")
        name = normalize_text_block(sheet.attrib.get("name", "Sheet"))
        if not rel_id or rel_id not in rel_map:
            continue
        target = posixpath.normpath(posixpath.join("xl", rel_map[rel_id]))
        sheets.append((name or "Sheet", target))
    return sheets


def _load_sheet_rows(root: ET.Element, shared_strings: list[str]) -> list[list[str]]:
    rows: list[list[str]] = []
    for row_node in root.findall(".//a:sheetData/a:row", _SHEET_NS):
        value_by_index: dict[int, str] = {}
        max_index = -1
        for cell in row_node.findall("a:c", _SHEET_NS):
            reference = cell.attrib.get("r", "")
            column_index = _column_letters_to_index("".join(char for char in reference if char.isalpha()))
            max_index = max(max_index, column_index)
            value_by_index[column_index] = _read_cell_value(cell, shared_strings)
        if max_index < 0:
            continue
        row = [value_by_index.get(index, "") for index in range(max_index + 1)]
        if any(row):
            rows.append(row)
    return rows


def _read_cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return normalize_text_block("".join(node.text or "" for node in cell.findall(".//a:t", _SHEET_NS)))

    raw_value = cell.findtext("a:v", default="", namespaces=_SHEET_NS)
    if cell_type == "s":
        try:
            return normalize_text_block(shared_strings[int(raw_value)])
        except (ValueError, IndexError):
            return ""
    if cell_type == "b":
        return "TRUE" if raw_value == "1" else "FALSE"
    return normalize_text_block(raw_value)


def _column_letters_to_index(column_letters: str) -> int:
    if not column_letters:
        return 0
    value = 0
    for character in column_letters.upper():
        value = value * 26 + (ord(character) - ord("A") + 1)
    return max(value - 1, 0)


def _rows_to_markdown(rows: list[list[str]]) -> str:
    max_columns = max(len(row) for row in rows)
    normalized_rows = [row + [""] * (max_columns - len(row)) for row in rows]
    header = normalized_rows[0]
    body = normalized_rows[1:] or [[""] * max_columns]

    markdown_lines = [
        f"| {' | '.join(_escape_markdown_cell(cell) for cell in header)} |",
        f"| {' | '.join('---' for _ in range(max_columns))} |",
    ]
    markdown_lines.extend(
        f"| {' | '.join(_escape_markdown_cell(cell) for cell in row)} |"
        for row in body
    )
    return "\n".join(markdown_lines)


def _escape_markdown_cell(value: str) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", " ").strip()
